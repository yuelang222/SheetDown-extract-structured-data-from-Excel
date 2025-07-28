import os
import openpyxl 
from collections import defaultdict
                                      
def extract(input_dir, output_dir="output"):
    # get input_dir base name
    excel_basename = os.path.splitext(os.path.basename(input_dir))[0]

    output_dir = os.path.join(output_dir, excel_basename)

    #create output directory
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    #create image directory
    img_dir = os.path.join(output_dir, "images")
    if not os.path.exists(img_dir):
        os.makedirs(img_dir)
    
    #load an Excel file using the openpyxl library
    wb = openpyxl.load_workbook(input_dir)
    content = ""
    
    #process each sheet in an Excel file
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]  
        content += f"# {sheet_name}\n\n"
        
        #process all merged cell ranges in a worksheet
        #using the top-left cell to represent each range and marking other cells as None in a dictionary.
        merged_cells = {}
        for merged_range in sheet.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col
            merged_cells[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if (r, c) != (min_row, min_col):
                        merged_cells[(r, c)] = None
        
        #process all images in a worksheet
        image_data = defaultdict(list)
        img_count = 0
        for img in sheet._images:
            img_count += 1
            #represent the position of the top-left corner of the image.
            row, col = img.anchor._from.row + 1, img.anchor._from.col + 1
            target_cell = (row, col)

            #find the merged cell containing the image
            found = False
            for r in range(row, 0, -1):
                for c in range(col, 0, -1):
                    if (r, c) in merged_cells and merged_cells[(r, c)] is not None:
                        r_span, c_span = merged_cells[(r, c)]
                        if r <= row < r + r_span and c <= col < c + c_span:
                            target_cell = (r, c)
                            found = True
                            break
                if found:
                    break

            img_filename = f"{sheet_name}_img_{img_count}.png"
            img_path = os.path.join(img_dir, img_filename)
            
            try:
                data = img.ref
                if hasattr(data, 'getvalue'):
                    data = data.getvalue()
                
                with open(img_path, "wb") as f:
                    f.write(data)
                
                #use html image tag
                image_data[target_cell].append(f"<img src='images/{img_filename}' alt='Image' style='display:inline; margin:0;'>")
                #use markdown image tag
                #image_data[target_cell].append(f"![Image](images/{img_filename})")
            except Exception as e:
                print(f"Image save failed: {e}")
        
        #generate html table
        max_row = sheet.max_row
        max_col = sheet.max_column
        content += "<table border=\"1\">"
        
        for row in range(1, max_row + 1):
            content += "<tr>"

            for col in range(1, max_col + 1):
                pos = (row, col)

                if pos in merged_cells and merged_cells[pos] is None:
                    continue

                cell = sheet.cell(row, col)
                value = cell.value
                cell_text = str(value).strip() if value is not None else ""
                
                rowspan = 1
                colspan = 1
                if pos in merged_cells and merged_cells[pos] is not None:
                    rowspan, colspan = merged_cells[pos]
                
                if pos in image_data:
                    if cell_text:
                        cell_text = f"{cell_text}<div style='margin-top:5px;'>"
                        cell_text += "".join(image_data[pos])
                        cell_text += "</div>"
                    else:
                        cell_text = "<div>" + "".join(image_data[pos]) + "</div>"

                td_attrs = []
                if rowspan > 1:
                    td_attrs.append(f"rowspan='{rowspan}'")
                if colspan > 1:
                    td_attrs.append(f"colspan='{colspan}'")
                
                attrs_str = " ".join(td_attrs)
                if attrs_str:
                    content += f"<td {attrs_str}>{cell_text}</td>"
                else:
                    content += f"<td>{cell_text}</td>"
            
            content += "</tr>"

        content += "</table>\n\n"
    
    output_file = os.path.join(output_dir, f"{excel_basename}.md")
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(content)
    
    return output_file


if __name__ == "__main__":
    input_dir = "example.xlsx"
    output_file = extract(input_dir)
    print(f"Markdown file saved to: {output_file}")
